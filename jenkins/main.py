from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Side, borders, Font, Alignment
import xml.etree.ElementTree as ET
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import pandas as pd


def psm_xml():
    xml_parameter = []
    root = ET.parse(r'./jenkins/config.xml').getroot()
    for child in root.iter('Parameter'):
        xml_parameter.append(child.attrib['obj'].strip())
    return xml_parameter


xml_dataframe = psm_xml()


df_xml = pd.DataFrame(xml_dataframe)

# print(df_xml)


def txt_extract():
    text_file = r"./jenkins/Psm_coverage_psm_dev.txt"
    words = {}
    with open(text_file, "r") as file:
        data = file.readlines()
        try:
            for index, line in enumerate(data):
                if index > 5 and line.split(" ")[0] != "--":
                    keys = line.split(" ")[0]
                    value = float(line.split(" ")[-1].strip().replace(",", "."))
                    if keys not in words.keys():
                        words[keys] = []
                        words[keys].append(value)
                    else:
                        words[keys].append(value)
        except:
            pass
    text_data = {}
    for coverage_key in words:
        coverage_value = sum(words[coverage_key])*100/len(words[coverage_key])
        text_data[coverage_key] = coverage_value

    return text_data


txt_dataframe = txt_extract()
df_txt = pd.DataFrame.from_dict(txt_dataframe, orient='index')
print(df_txt)


def xl_parameters():
    wb = load_workbook(r"./jenkins/Individual_parameter_coverage.xlsx")
    ws = wb["Sheet1"]

    start_row = 0
    start_col = 0

    for row in ws:
        for cell in row:
            if cell.value == "Parameter Name":
                start_row = cell.row
                start_col = cell.column

    parameter_name = {}
    for row in range(start_row + 1, ws.max_row):
        s = ws.cell(row=row, column=start_col)
        parameter_name[s.value] = ws.cell(row=row, column=start_col+1).value
    del parameter_name[None]
    return parameter_name


excel_dataframe = xl_parameters()
df_excel = pd.DataFrame.from_dict(excel_dataframe, orient='index')
print(df_excel)


def missing_parameter():
    missing_parameters = []
    excel = xl_parameters()
    xml = psm_xml()
    for i in excel.keys():
        if i not in xml:
            missing_parameters.append(i)
    return missing_parameters


def coverage_results():

    result = []
    for i in txt_extract().keys():
        if i in xl_parameters().keys():
            if txt_extract()[i] > xl_parameters()[i]:
                result.append((i, xl_parameters()[i], txt_extract()[i]))

    return result


def xl_writing(sheet_name, values):
    side = Side(border_style="thick")
    border = borders.Border(
        left=side,
        right=side,
        top=side,
        bottom=side,
    )

    wb2 = Workbook()
    ws2 = wb2.create_sheet(sheet_name, 0)

    ws2["B2"] = sheet_name
    ws2["B2"].font = Font(color="ffffff", size="18", bold=True)
    ws2["B2"].fill = PatternFill("solid", start_color="0458AB")
    ws2['B2'].alignment = Alignment(horizontal='center', vertical='center')

    ws2['B3'] = "Parameter Name"
    ws2['C3'] = "Coverage Ratio(%)"

    if sheet_name == "Missing Parameters":
        ws2.merge_cells("B2:C2")
        s = values
        for row in range(0, len(s)):
            for col in range(1, 2):
                ws2[chr(66)+str(row + 4)] = s[row]
                ws2[chr(67) + str(row + 4)] = xl_parameters()[s[row]]
    else:
        ws2.merge_cells("B2:D2")
        ws2.merge_cells("C3:D3")
        ws2.merge_cells("B3:B4")

        s = values

        for row in range(0, len(s)):
            for col in range(1, 3):
                ws2[chr(66) + str(row + 5)] = s[row][0]
                ws2[chr(67) + str(row + 5)] = s[row][1]
                ws2[chr(68) + str(row + 5)] = s[row][2]

    ws2.column_dimensions["B"].width, ws2.column_dimensions["C"].width, ws2.column_dimensions["D"].width = 42, 18, 15

    for row in range(1, ws2.max_row+1):
        for col in range(1, ws2.max_column+1):
            if ws2.cell(row=row, column=col).value:
                ws2.cell(row=row, column=col).border = border

    wb2.save(r"./jenkins/Result_Parameters.xlsx")


def send_email():
    fromaddr = "saisaran796@gmail.com"
    toaddr = "saisaran796@gmail.com"

    msg = MIMEMultipart()

    msg['From'] = fromaddr

    msg['To'] = toaddr

    msg['Subject'] = "Subject of the Mail"

    body = "Body_of_the_mail"

    msg.attach(MIMEText(body, 'plain'))

    filename = "Result_Parameters.xlsx"
    attachment = open(r"./jenkins/Result_Parameters.xlsx", "rb")

    p = MIMEBase('application', 'octet-stream')

    p.set_payload(attachment.read())

    encoders.encode_base64(p)

    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)

    msg.attach(p)

    s = smtplib.SMTP('smtp.gmail.com', 587)

    s.starttls()

    s.login(fromaddr, "Romanking@123")

    text = msg.as_string()

    s.sendmail(fromaddr, toaddr, text)

    s.quit()


if len(missing_parameter()) != 0:
    name = "Missing Parameters"
    result = missing_parameter()
    xl_writing(name, result)
    send_email()
else:
    name = "Coverage Results"
    result = coverage_results()
    xl_writing(name, result)
    if len(txt_extract()) > len(result):
        pass
        send_email()
